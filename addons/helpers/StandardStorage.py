from datetime import datetime

import numpy as np

import openpyxl
import timeit


def create_result_storage(models_instances, metric_txt, prediction_lines):
    if not isinstance(metric_txt, list):
        metric_txt = [metric_txt]

    results = StandardStorage()
    results.append_columns(['timestamp', 'y_true'])
    first_model_index = None
    columns_per_model = None
    for index, inst in enumerate(models_instances):
        columns = [*[str(inst) + str(index) + "_" + m for m in metric_txt],
                   *[str(inst) + str(index) + "_" + m + "_trend" for m in metric_txt],
                   *[str(inst) + str(index) + "_" + str(x) for x in range(prediction_lines)]]
        if columns_per_model is None:
            columns_per_model = len(columns)

        if first_model_index is None:
            first_model_index = columns[0]

        results.append_columns(columns)

    return results, results.get_column_index(first_model_index), columns_per_model

'''
    Function reads csv and rewrite it into standard storage. All apart from last column are features. Last column is 
    corresponding label
'''
def read_csv_database_file(database_file_path, delimiter=",", encoding='utf-8', converters=None, skip_header=0):
    try:
        data = np.genfromtxt(database_file_path, skip_header=skip_header, delimiter=delimiter, dtype=None,
                             encoding=encoding, converters=converters)
        # first column is timestamp next are additional arguments. The last one is Y
        #_data_X = data[:, 0:-1]
        _storage = StandardStorage()
        _storage.headers = ["X" + str(i) for i in list(range(len(data[0])))]
        _storage.init_data(data)
        return _storage

    except OSError:
        print("The given file", "\"" + database_file_path + "\"", "doesnt exist")
        print("Make correction and try again")
        raise OSError()

class StandardStorageIterator():
    def __init__(self, standard_storage):
        self._currentIndex = 0
        self._standard_storage = standard_storage
        pass

    def __iter__(self):
        return self
    def __next__(self):
        if self._currentIndex < len(self._standard_storage):
            self._currentIndex = self._currentIndex + 1
            return self._standard_storage.data[self._currentIndex - 1]

        raise StopIteration

'''
    The 
'''
class StandardStorage():
    def __init__(self, file_path=None):
        self._parameters = {}
        self._rows = 0
        self._excel_file = openpyxl.Workbook()
        self._excel_sheet = self._excel_file.active

        # control value
        self._excel_sheet['A1'] = "#001001"

        self._headers = []
        self._data = None
        self._path = None

        if file_path is not None:
            self.load_excel_format(file_path)

    @property
    def file_path(self):
        return self._path

    @property
    def parameters(self):
        return self._parameters

    @parameters.setter
    def parameters(self, val):
        self._parameters = val

    @property
    def headers(self):
        return self._headers

    @headers.setter
    def headers(self, val):

        if self._data is not None:
            new_data = np.zeros((self._data.shape[0], len(val)))

            for index,h in enumerate(val):
                # if new list is
                for j, hh in enumerate(self._headers):
                    if hh == h:
                        # swap column
                        new_data[:, index] = self.data[:,j]
            self._data = new_data
        self._headers = val

    @property
    def data(self):
        return self._data

    @data.setter
    def data(self, val):
        self._data = val

    def init_data(self, data):
        self._data = data.copy()

    def init_zeros(self, length):
        self._data = np.zeros((length,len(self._headers)))
        self._data[:] = np.nan

    def append_columns(self, columns):
        self.headers = [*self.headers, *columns]

    def remove_column(self, column):
        try:
            headers = self.headers
            headers.remove(column)
            self.headers = headers
        except ValueError:
            pass

    def remove_columns(self, columns):
        try:
            headers = self.headers
            for column in columns:
                headers.remove(column)
            self.headers = headers
        except ValueError:
            pass

    def append(self, row):
        # if data exists and columns count equals headers
        if self._data is not None and len(self._data.shape) == 2 and self._data.shape[1] == len(self.headers):
            self._data = np.append(self._data, np.array(row).reshape((1,len(row))), axis=0)
        else:
            self._data = np.zeros((1, len(self._headers)), dtype=object)
            self._data[0] = np.array(row)

    def _row(self, index):
        return self._data[index, :]

    def _column(self, column):
        return self._data[:, column]

    def get_column_index(self, header):
        return self._headers.index(header)

    def get_column(self, header):
        index = self.get_column_index(header)
        if index >= 0:
            return self._data[:, index]

    def get_column_by_index(self, index):
        return self._data[:, index]

    def get_row(self, value):
        return self._data[value, :]

    def get(self, header=None, indexes=None):
        if header is None and indexes is None:
            return self._data
        elif indexes is None:
            return self.get_column(header)
        elif header is None:
            return self.get_row(indexes)
        else:
            index = self.get_column_index(header)
            if index >= 0:
                return self._data[indexes, index]

    @staticmethod
    def _column_name(base, shift):
        tb = ord(base)
        base_shift = tb - ord('A')

        full_steps = int((shift + base_shift) / (ord('Z')+1 - ord('A')))
        mod_step = int((shift + base_shift) % (ord('Z')+1 - ord('A')))

        if full_steps == 0:
            return chr(ord('A') + mod_step)
        else:
            return str(chr(ord('A') + full_steps-1)) + str(chr(ord('A') + mod_step))

    @staticmethod
    def _cell_name(base, shift, index):
        return StandardStorage._column_name(base,shift) + str(index)

    @staticmethod
    def timestamp_as_name():
        dt = datetime.now()
        return dt.strftime("%Y%m%d%H%M%S")

    '''
        Function allow to set value in a particular cell
    '''
    def set(self, header,indexes,values):
        index = self.get_column_index(header)
        self._data[indexes, index] = values
    '''
        Function allows to save data in excel instead of 
    '''
    def save_excel_format(self, path):
        for index, key in enumerate(self._parameters):
            row = str(index+3)
            self._excel_sheet["A" + row] = str(key)
            self._excel_sheet["B" + row] = str(self._parameters[key])

        tb = 'C'

        for index in range(len(self.headers)):
            self._excel_sheet[self._cell_name(tb, index, 1)] = self.headers[index]

        if self._data is not None:
            self._excel_sheet["A2"] = self.data.shape[0]
            self._excel_sheet["B2"] = self.data.shape[1]
            for index in range(1,self.data.shape[0]+1):
                for j in range(0,self.data.shape[1]):
                    self._excel_sheet[self._cell_name(tb, j, index+1)] = self.data[index-1,j]
        else:
            self._excel_sheet["A2"] = 0
            self._excel_sheet["B2"] = len(self.headers)

        print(self._excel_file)
        self._excel_file.save(path)

    def save_csv_format(self, path):
        with open(path, "w") as file:
            file.write(";".join(self.headers) + "\n")
            if self._data is not None:
                for index in range(0, self.data.shape[0]):
                    file.write(";".join(["{0:.3f}".format(self.data[index, k]) for k,_ in enumerate(self.data[index])]) + "\n")
    '''
        Load excel file from disc
    '''
    def load_excel_format(self, path):
        self._path = path
        self._excel_file = openpyxl.load_workbook(path)
        self._excel_sheet = self._excel_file.active
        self._rows = self._excel_sheet.max_row - 1

        if self._excel_sheet['A1'].value != "#001001":
            raise Exception("Incorrect file format!")

        index = 3
        while True:
            row = str(index)
            key = self._excel_sheet["A" + row].value
            value = self._excel_sheet["B" + row].value

            if key is None or value is None:
                break

            self._parameters[key] = value
            index = index + 1

        columns = int(self._excel_sheet['B2'].value)
        rows = int(self._excel_sheet['A2'].value)

        tb = 'C'

        self.headers = [[x.value for x in y] for y in self._excel_sheet[self._cell_name(tb, 0, 1):self._cell_name(tb, columns-1, 1)]][0]
        self.data = np.array([[x.value for x in y] for y in self._excel_sheet[self._cell_name(tb, 0, 2):self._cell_name(tb, columns-1, rows+2-1)]], dtype=object)

        return self

    def __str__(self):
        return "Params:\n" + "\n".join([" " + str(key) + ": " + str(self._parameters[key]) for key in self._parameters]) + "\nH:" + \
               "|".join(self._headers) + "\n" + str(self._data)

    def __len__(self):
        if self._data is not None:
            return self._data.shape[0]
        else:
            return 0

    def __iter__(self):
        return StandardStorageIterator(self)

    def __getitem__(self, header, index):
        return self.get(header, index)

    def __add__(self, other):
        if self.headers != other.headers:
            raise ValueError("Cannot add StandardStorage with different columns")

        storage = StandardStorage()
        storage.headers = self.headers
        storage.init_zeros(len(self) + len(other))

        storage.data[:len(self)] = self._data
        storage.data[len(self):] = other.data

        return storage

        # print(header, index)
        # return self.get(header, index)


# Test cases

if __name__ == "__main__":
    print("Test cases")

    class TestObj():
        def __init__(self):
            self.var = "var"

    ss = StandardStorage()
    ss.parameters["param_int_1"] = 123
    ss.parameters["param_string_1"] = "Ala ma kota"
    ss.parameters["param_object_1"] = TestObj()

    ss.headers = ["DT", "Production", "Consumption"]
    ss.append([1, 25, 0.74])
    ss.append([1, 25, 0.74])
    ss.append([1, 25, 0.74])
    ss.append([1, 25, 0.74])
    ss.append([1, 25, 0.74])
    ss.append([1, 25, 0.75])

    # append some data columns
    ss.append_columns(["+" + str(x) + " prod" for x in range(1, 5)])

    # append some metrics column
    ss.append_columns(["R2", "MSE"])

    # load indexes of some columns
    index_first = ss.get_column_index("+1 prod")
    index_last = ss.get_column_index("+3 prod")

    # row iteration
    for i,element in enumerate(ss):
        # load data from row (input data)
        args = element[0:3]

        # write arrays (result from prediction)
        element[index_first:index_last+1] = args*2

        # write particular cells (metrics, computation time or others)
        ss.set("R2", i, -1)
        ss.set("MSE", i, 22)

    #take a whole array
    print(ss.get())

    # take a column as a row
    print(ss.get_column("R2"))
    print(ss.get("R2"))

    #take a row
    print(ss.get_row(0))
    print(ss.get(None, 0))

    # print object
    print(ss)

    # save file on the disc
    ss.save_excel_format("temp/temp.xls")